from server.settings import BASE_DIR
from apps.audit.models import Atask, AtaskItem, AtaskIssue, AtaskProblem
from openpyxl import load_workbook
import os
from apps.audit.models import R_LEVEL_DICT, Company
from docxtpl import DocxTemplate
import logging
import re
from xml.sax.saxutils import escape

myLogger = logging.getLogger('log')

templ = os.path.join(BASE_DIR, "media/muban/xxxx任务问题清单.xlsx")
def export_issue_excel(atask:Atask):
    wb = load_workbook(filename=templ)
    ws = wb.active
    ws["A1"] = f'{atask.company.name}安全审计问题清单'
    ws["C2"] = f"{atask.company.audit_scope}"
    member_name_list = list(atask.team.values_list("member__name", flat=True))
    ws["C3"] = ','.join(member_name_list)
    kfx_qs = AtaskItem.objects.filter(atask=atask, is_suit=False).values("standarditem__number", "standarditem__full_score")
    kfx_number_list = [kfx["standarditem__number"] for kfx in kfx_qs]
    kfx_full_score_list = [kfx["standarditem__full_score"] for kfx in kfx_qs]
    kfx_str = '/'.join(kfx_number_list)
    kfx_score_str = '+'.join([str(score) for score in kfx_full_score_list])
    ws["C4"] = f'{kfx_str}={kfx_score_str}={sum(kfx_full_score_list)}'
    count = AtaskIssue.objects.filter(atask=atask).count()
    ws.move_range("A9:F12", rows=count)
    # ws.insert_rows(6, count)
    # kill_score_all = 0
    # for ind, item in enumerate(AtaskIssue.objects.filter(atask=atask)):
    #     row = 6+ind
    #     standarditem = item.standarditem
    #     if standarditem:
    #         if standarditem.level == 10:
    #             ws[f"A{row}"] = standarditem.content
    #         elif standarditem.level == 20:
    #             ws[f"A{row}"] = standarditem.parent.content
    #         elif standarditem.level == 30:
    #             ws[f"A{row}"] = standarditem.parent.parent.content
    #     ws[f"B{row}"] = standarditem.number if standarditem else ""
    #     ws[f"C{row}"] = item.content
    #     ws[f"D{row}"] = R_LEVEL_DICT.get(item.risk_level, "")
    #     ws[f"E{row}"] = item.kill_score if item.kill_score else ""
    #     if item.kill_score is not None:
    #         kill_score_all += item.kill_score
    #     ws[f"F{row}"] = item.create_by.name
    path = f'/media/temp/任务问题清单_{atask.id}.xlsx'
    wb.save(BASE_DIR + path)
    return path




def export_issue_docx(atask:Atask, type=1):
    templ2 = os.path.join(BASE_DIR, "media/muban/xxx任务问题清单.docx")
    if str(type) == "2":
        templ2 = os.path.join(BASE_DIR, "media/muban/ddd任务问题清单.docx")
    doc = DocxTemplate(templ2)
    title = f'{atask.company.name}安全审计问题清单'
    audit_scope = atask.company.audit_scope if atask.company.audit_scope else ""
    member_name_list = list(atask.team.values_list("member__name", flat=True))
    member_names = ','.join(member_name_list)
    kfx_qs = AtaskItem.objects.filter(atask=atask, is_suit=False).values("standarditem__number", "standarditem__full_score")
    kfx_number_list = [kfx["standarditem__number"] for kfx in kfx_qs]
    kfx_full_score_list = [kfx["standarditem__full_score"] for kfx in kfx_qs]
    kfx_str = '/'.join(kfx_number_list)
    kfx_score_str = '+'.join([str(score) for score in kfx_full_score_list])
    sum_kfx_score = sum(kfx_full_score_list)
    if sum_kfx_score == 0:
        kfx = '无'
    else:
        kfx = f'{kfx_str}={kfx_score_str}={sum_kfx_score}'
    issues = []
    kill_score_all = 0
    for item in AtaskIssue.objects.filter(atask=atask).order_by("atask", "standarditem__number_sort", "-create_time"):
        standarditem = item.standarditem
        if standarditem:
            if standarditem.level == 10:
                content = standarditem.content
            elif standarditem.level == 20:
                content = standarditem.parent.content
            elif standarditem.level == 30:
                content = standarditem.parent.parent.content
        else:
            content = ""
        issues.append({
            "id": item.id,
            "c": content,
            "n": standarditem.number if standarditem else "",
            # "issue_content": item.content if item.content else "",
            "issue_content": escape(item.content) if item.content else "",
            "r": R_LEVEL_DICT.get(item.risk_level, ""),
            "k": item.kill_score if item.kill_score is not None else "",
            "x": item.create_by.name if item.create_by.name else "",
        })
        if item.kill_score is not None:
            kill_score_all += item.kill_score
        
    problems = []
    for ind, item in enumerate(AtaskProblem.objects.filter(atask=atask).order_by("atask", "-create_time")):
        problems.append({
            "index": ind+1,
            "content": item.content if item.content else "",
        })
    context = {
        "title": title,
        "audit_scope": audit_scope,
        "member_names": member_names,
        "kfx": kfx,
        "issues": issues,
        "problems": problems,
        "s": kill_score_all
    }
    
    doc.render(context)
    path = f'/media/temp/任务问题清单_{atask.id}_{type}.docx'
    doc.save(BASE_DIR + path)
    return path


def deep_clean(data):
    if isinstance(data, dict):
        return {k: deep_clean(v) for k, v in data.items()}
    elif isinstance(data, list):
        return [deep_clean(item) for item in data]
    elif isinstance(data, str):
        # 移除非法控制字符（保留 \t, \n, \r）
        data = re.sub(r'[\x00-\x08\x0B\x0C\x0E-\x1F]', '', data)
        # 不转义 XML，docxtpl 会自动处理
        return data
    else:
        return data  # 保持数字、布尔值等不变

def export_atask_report(year: int, company: str):
    """导出审计报告"""
    templ = os.path.join(BASE_DIR, "media/muban/XXXX有限公司安全审计报告.docx")
    company = Company.objects.get(id=company)
    atasks = Atask.objects.filter(year=year, company__id=company)
    if atasks.count() == 1:
        return ""
    data = {}
    data["company_name"] = atask.company.name
    data["dates"] = f'{atask.start_date.strftime("%Y年%m月%d日")}至{atask.end_date.strftime("%Y年%m月%d日")}' if atask.start_date and atask.end_date else ""
    data["leader_name"] = atask.leader.name
    data["member_name_list"] = list(atask.team.values_list("member__name", flat=True))
    data["member_name_list"].remove(atask.leader.name)
    data["member_names"] = ','.join(data["member_name_list"])
    data["member_count"] = len(data["member_name_list"]) + 1
    data["days"] = (atask.end_date - atask.start_date).days + 1
    
    docx = DocxTemplate(templ)
    docx.render(data)
    path = f'/media/temp/{atask.year}_{atask.company.name}_安全审计报告.docx'
    docx.save(BASE_DIR + path)
    return path
    
