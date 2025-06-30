from apps.audit.models import Standard, StandardItem, Atask, AtaskIssue
from openpyxl import load_workbook
from rest_framework.exceptions import ParseError
from django.core.mail import EmailMessage
from apps.system.models import User
import re

def get_number_sort(number: str):
    parts = str(number).split('.')
    return '.'.join([f"{int(part):03d}" for part in parts])

def daoru_standard(path: str, sta: Standard):
    wb = load_workbook(path)
    current_item1 = None
    current_item2 = None
    total_score = 0
    ws = wb["审计标准"]
    current_item1 = None
    current_item2 = None
    current_item3 = None
    current_item_2_chid = []
    share_method = None
    current_item1_score = 0
    current_item2_score = 0
    i = 3
    while ws[f'g{i}'].value:
        number1 = ws[f'a{i}'].value
        content1 = ws[f'b{i}'].value
        number2 = ws[f'c{i}'].value
        content2 = ws[f'd{i}'].value
        rlevel_str = ws[f'e{i}'].value.strip() if ws[f'e{i}'].value else None
        rlevel = 20 if rlevel_str == "★" else 30 if rlevel_str == "★★" else 40 if rlevel_str == "★★★"  else 10
        number3 = ws[f'f{i}'].value
        content3 = ws[f'g{i}'].value

        method = ws[f'h{i}'].value
        share_method = method if method else share_method
        full_score = ws[f'i{i}'].value
        if number1:
            x1, _ = StandardItem.objects.get_or_create(
                number = number1,
                standard = sta,
                defaults={
                    "content": content1,
                    "level": 10,
                }
            )
            if x1.number_sort is None:
                x1.number_sort = get_number_sort(number1)
                x1.save(update_fields=["number_sort"])
            if current_item1:
                current_item1.full_score = current_item1_score
                current_item1.save()
                current_item1_score = 0
            current_item1 = x1
        if number2:
            x2, _ = StandardItem.objects.get_or_create(
                number = number2,
                standard = sta,
                defaults={
                    "content": content2,
                    "parent": current_item1,
                    "level": 20,
                    "method": share_method
                }
            )
            if x2.number_sort is None:
                x2.number_sort = get_number_sort(number2)
                x2.save(update_fields=["number_sort"])
            if current_item2:
                current_item2.full_score = current_item2_score
                current_item2.save()
                current_item2_score = 0
                current_item_2_chid = []
            current_item2 = x2
        if number3:
            current_item3, is_create = StandardItem.objects.get_or_create(
                    number = number3,
                    standard = sta,
                    defaults={
                        "content": content3,
                        "parent": current_item2,
                        "level": 30,
                        "risk_level": rlevel,
                        "method": method,
                        "full_score": full_score
                    }
                )
            if current_item3.number_sort is None:
                current_item3.number_sort = get_number_sort(number3)
                current_item3.save(update_fields=["number_sort"])
            if full_score:
                current_item2_score += full_score
                current_item1_score += full_score
                total_score += full_score
            if full_score is None:
                current_item3.is_concern = False
                current_item3.save(update_fields=["is_concern"])
                if current_item2.is_concern is False:
                    current_item2.is_concern = True
                    current_item2.save(update_fields=["is_concern"])
            else:
                current_item3.is_concern = True
                current_item3.save(update_fields=["is_concern"])

            current_item_2_chid.append(current_item3)
            if len(current_item_2_chid) == 2:
                if full_score is None:
                    current_item_2_chid[0].full_score = None
                    current_item_2_chid[0].is_concern = False
                if method is None:
                    current_item_2_chid[0].method = None
                current_item_2_chid[0].save(update_fields=["full_score", "method", "is_concern"])
        if content3 and number3 is None:
            if content3 in current_item3.content:
                current_item3.content = content3
            else:
                current_item3.content = current_item3.content + ";" + content3
            current_item3.save(update_fields=["content"])
        i += 1
    if current_item2:
        current_item2.full_score = current_item2_score
        current_item2.save()
    if current_item1:
        current_item1.full_score = current_item1_score
        current_item1.save()
    sta.total_score = total_score
    sta.save()
    return sta

def daoru_issue(path:str, atask: Atask, user):
    standarditems = StandardItem.objects.filter(standard=atask.standard, level=30)
    st_dict = {}
    for item in standarditems:
        st_dict[item.number] = item

    wb = load_workbook(path)
    ws = wb["Sheet1"]
    i = 3
    cal_sitem = []
    while ws[f'b{i}'].value:
        standarditem:StandardItem = st_dict.get(ws[f'b{i}'].value, None)
        cal_sitem.append(standarditem)
        number = ws[f'a{i}'].value
        content = ws[f'c{i}'].value
        kill_score = int(ws[f'd{i}'].value)
        if not number:
            raise ParseError(f"{i}行-问题编号不存在")
        if standarditem is None:
            raise ParseError(f"{i}行-标准项不存在")
        AtaskIssue.objects.get_or_create(
            number = number,
            standarditem = standarditem,
            atask = atask,
            defaults={
                "content": content,
                "kill_score": kill_score,
                "create_by": user
            }
        )
        i = i + 1

    if cal_sitem:
        for item in cal_sitem:
            AtaskIssue.cal(atask, user, item)


def sendMail(atask:Atask):
    team = atask.team
    to = []
    content = atask.notify_content
    if content:
        pass
    else:
        raise ParseError("通知内容不能为空")
    def is_valid_email(email):
        pattern = r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'
        return bool(re.match(pattern, email))
    
    for m in team:
       user:User = m.member
       mail = user.username
       if is_valid_email(mail):
           to.append(mail)
    if is_valid_email(atask.company.email):
        to.append(atask.company.email)

    email = EmailMessage(
        subject=f"【{atask.company.name}】审核任务",
        body=content,
        to=to
    )
    email.send()