import os
from django.conf import settings
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter
from docxtpl import DocxTemplate
from PIL import Image as PILImage
from io import BytesIO


def export_docx(template_path: str, context_data: dict, file_name: str):
    """
    Word导出
    :param template_path: 模板路径
    :param context_data: 数据
    :param file_name: 保存的文件名
    :return:返回文件路径
    """
    docx = DocxTemplate(settings.BASE_DIR + template_path)
    docx.render(context_data)
    file_path = f'/media/temp/{file_name}'
    save_path = settings.BASE_DIR + file_path
    docx.save(save_path)
    return file_path


def len_byte(value):
    # 获取字符串长度，一个中文的长度为2
    length = len(value)
    utf8_length = len(value.encode('utf-8'))
    length = (utf8_length - length) / 2 + length
    return int(length)


def export_excel(field_data: list, data: list, FileName: str, img_field_index: list = []):
    """
    Excel导出（支持图片）
    :param field_data: 表头数据，可以是字符串列表或字典列表 [{'name':'', 'type':''}] 或 ['字段1', '字段2']
    :param data: 数据源
    :param FileName: 文件保存名字
    :param img_field_index: 图片字段索引列表（从0开始）
    :return: 返回文件的下载url完整路径
    """
    wb = Workbook()
    ws = wb.active
    ws.title = 'Sheet1'
    
    # 样式设置
    header_fill = PatternFill(start_color="003C8D", end_color="003C8D", fill_type="solid")
    header_font = Font(name='楷体', size=10, bold=True, color="FFFFFF")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # 边框样式
    thin_border = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000")
    )
    
    content_font = Font(name='楷体', size=10)
    content_alignment = Alignment(horizontal="center", vertical="center")
    
    # 处理表头数据格式
    header_fields = []
    for i, field in enumerate(field_data):
        if isinstance(field, dict):
            header_fields.append({
                'name': field.get('name', ''),
                'type': field.get('type', 'text'),
                'index': i
            })
        else:
            header_fields.append({
                'name': str(field),
                'type': 'img' if i in img_field_index else 'text',
                'index': i
            })
    
    # 写入表头
    for col_idx, field in enumerate(header_fields, 1):
        cell = ws.cell(row=1, column=col_idx, value=field['name'])
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border
        
        # 设置列宽
        letter = get_column_letter(col_idx)
        if field['type'] == 'img':
            ws.column_dimensions[letter].width = 15  # 图片列宽
        else:
            ws.column_dimensions[letter].width = 14  # 默认列宽
    
    # 确定栏位宽度
    col_widths = [len(str(field['name'])) for field in header_fields]
    
    def process_image(img_path):
        """处理单张图片并返回BytesIO"""
        # 获取文件大小
        file_size_kb = os.path.getsize(img_path) / 1024
        
        with PILImage.open(img_path) as img:
            # 如果文件小于等于50KB，直接返回原图片
            if file_size_kb <= 50:
                buffer = BytesIO()
                img.save(buffer, format=img.format if img.format else "JPEG")
                buffer.seek(0)
                return buffer
            else:
                # 限制最大尺寸
                img.thumbnail((1600, 1600))
                buffer = BytesIO()
                if img.mode == 'RGBA':
                    img = img.convert('RGB')
                img.save(buffer, format="JPEG", quality=85)
                buffer.seek(0)   
                return buffer
        
    for row_idx, row_data in enumerate(data, 2):
        for col_idx, cell_value in enumerate(row_data):
            if col_idx < len(header_fields):
                field_type = header_fields[col_idx]['type']
                cell = ws.cell(row=row_idx, column=col_idx + 1)
                
                if cell_value is None:
                    cell_value = ''
                
                # 处理图片
                if field_type == 'img' and cell_value:
                    try:
                        img_path = settings.BASE_DIR + str(cell_value)
                        if os.path.exists(img_path):
                            img_buffer = process_image(img_path)
                            img = Image(img_buffer)
                            # 设置图片尺寸
                            img.width = 90
                            img.height = 90
                            ws.row_dimensions[row_idx].height = 70
                            ws.add_image(img, get_column_letter(col_idx + 1) + str(row_idx))
                            cell.value = ''  # 清空单元格值，只显示图片
                        else:
                            cell.value = '图片不存在'
                    except Exception as e:
                        cell.value = f'图片加载错误: {str(e)}'
                else:
                    cell.value = str(cell_value) if cell_value is not None else ''
                
                # 设置内容样式
                cell.font = content_font
                cell.alignment = content_alignment
                cell.border = thin_border
                
                # 更新列宽
                cell_len = len(str(cell_value)) if cell_value else 0
                if cell_len > col_widths[col_idx]:
                    col_widths[col_idx] = cell_len
    
    # 调整列宽
    for col_idx, width in enumerate(col_widths, 1):
        letter = get_column_letter(col_idx)
        adjusted_width = min(max(width, 10), 36) + 2  # 限制宽度范围并增加边距
        ws.column_dimensions[letter].width = adjusted_width
    
    # 保存文件
    FileNameF = FileName + datetime.now().strftime('%Y%m%d%H%M%S') + '.xlsx'
    path = '/media/temp/'
    pathRoot = settings.BASE_DIR + path
    
    if not os.path.exists(pathRoot):
        os.makedirs(pathRoot)
    
    path_name = os.path.join(pathRoot, FileNameF)
    wb.save(path_name)
    
    return path + FileNameF